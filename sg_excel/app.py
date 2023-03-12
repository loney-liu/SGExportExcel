import pprint
from openpyxl import Workbook
from openpyxl.utils.cell import (cols_from_range,
                                 rows_from_range,
                                 column_index_from_string,
                                 get_column_letter, 
                                 coordinate_to_tuple)
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import os
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import logging
from shotgun_api3 import Shotgun

# Change this if the excel need to be saved to other path. 
# Otherwise, it is stored in the directory of the AMI script.
folder_path = os.path.realpath(os.path.dirname(__file__))
img_temp = ("{0}/imgs".format(folder_path))

sg_log = logging.getLogger('shotgun_api3')

logging.basicConfig(filename=("{0}/{1}.log").format(folder_path, os.path.splitext(os.path.basename(__file__))[0]), 
                    format='%(asctime)s %(message)s',
                    filemode='w')

logger = logging.getLogger()

# Setting the threshold of logger to DEBUG
logger.setLevel(logging.DEBUG)

if not os.path.exists(img_temp):
    os.mkdir(img_temp)
    logger.info(("Create image temp folder: {0}").format(img_temp))

def export(url):
    # Parse URL data
    parts = urlparse(url)
    directories = parts.path.strip('/').split('/')
    queries = parts.query.strip('&').split('&')
    url_data = parse_qs(parts.query)
    
    # Get ShotGrid query data from url data
    ids = url_data['ids'][0].split(',')
    # selected_ids = url_data['selected_ids']
    cols = url_data['cols']
    column_display_names = url_data['column_display_names']
    scritp_user = directories[0]
    scritp_key = directories[1]
    user_login = url_data['user_login'][0]
    site_url = "https://" + url_data['server_hostname'][0]
    entity_type = url_data['entity_type'][0]

    export_excel=ExportExcel(url, site_url, scritp_user, scritp_key, user_login, cols, column_display_names, entity_type, ids)
    export_excel.export_excel()

class ExportExcel():
    def __init__(self, url, site_url, scritp_user, scritp_key,user_login, cols, column_display_names, entity_type, ids):
        self.__create_url_log(url)
        self.__first_row = 1
        self.__first_col = 1
        
        # Default image size in Excel
        self.__image_size = (160, 90)
        
        # ShotGrid information
        self.__site_url = site_url
        self.__scritp_user= scritp_user
        self.__scritp_key = scritp_key
        self.__user_login = user_login
        self.__cols = cols
        self.__column_display_names = column_display_names
        self.__ids = ids
        self.__entity_type = entity_type
        self.__sg = self.__create_connection()

        self.__col_json = self.__map_code_name()
        self.__sg_data = self.__query_sg_data()

        now_time = datetime.now().strftime('%Y-%m-%d_%H%M%S') 
        self.__excel_file = ('{0}/{1}_{2}.xlsx').format(folder_path, self.__entity_type, now_time)

    def __create_url_log(self, url):
        # fh = open(('{0}/echo.log').format(folder_path), 'w')
        # fh.write(pprint.pformat((url)))
        # fh.close()
        logger.info(("URL: '{0}'").format(url))

    def __query_sg_data(self):
        '''
        Query ShotGrid data
        '''
        try:
            query_result = self.__sg.find(self.__entity_type,[["id","in",[eval(i) for i in  self.__ids]]],self.__cols)
            logger.info(("Find: {0}").format(query_result))
            return query_result
        except Exception as e:
            logger.info(("Find Error: {0}").format(e))

    def __map_code_name(self):
        '''
        create dictionary for code and display_name
        '''
        col_json = {self.__cols[i]:self.__column_display_names[i] for i in range(len(self.__cols))}
        logger.info(("Converted: {0}").format(col_json))
        return col_json

    def __create_connection(self):
        try:
            sg = Shotgun(self.__site_url, self.__scritp_user, self.__scritp_key, sudo_as_login=self.__user_login)

            logger.info(("Connected"))
            return sg
        except Exception as e:
            logger.info(("Connection Error: {0}").format(e))

    def __sg_2_excel(self):
        '''
        Convert sg find data to excel data.
        '''
        excel_data = []
        try:
            for data in self.__sg_data:
                e={}
                for key, value in data.items():
                    if key=='type':
                        continue
                    if key.__contains__('step'):
                        continue
                    if key=='image':
                        if value is not None:
                            self.__download_thumbnail(data['id'])
                        continue

                    if type(value) == type({}):
                        value = value.get('name')

                    if type(value) == type([]):
                        cv = []
                        if len(value) == 0:
                            value = None
                        else: 
                            for a in value:
                                cv.append(a.get('name'))
                        value = ', '.join([str(elem) for elem in cv])

                    if self.__col_json.get(key):
                        e[self.__col_json[key]]=value
                    else:
                        e[key]=value
                if e:
                    excel_data.append(e)
                
            logger.info(("Parse SG Data: {0}").format(excel_data))
        except Exception as e:
            logger.info(("Parse SG Data Error: {0}").format(e))
        
        return excel_data

    def __download_thumbnail(self, entity_id):
        local_file_path =('{0}/{1}.jpg').format(img_temp, entity_id)
        if not os.path.exists(local_file_path):
            try:
                d = self.__sg.download_attachment({'url': ('{0}/thumbnail/full/{1}/{2}').format(self.__site_url, self.__entity_type, entity_id)}, local_file_path)
                logger.info(("Downloaded: {0}").format(local_file_path))
            except Exception as e:
                logger.info(("Download Error: {0}").format(e))
                if os.path.exists(local_file_path):
                    os.remove(local_file_path)
                    logger.info(("Existed: {0}").format(local_file_path))
        else:
            logger.info(("Existed: {0}").format(local_file_path))
            

    def export_excel(self):
        try:
            workbook = Workbook()
            workbook.sheetnames
            sheet = workbook.active
            sheet.title = self.__entity_type
            excel_data = self.__sg_2_excel()

            # Set width of column A
            sheet.column_dimensions[get_column_letter(self.__first_col)].width = self.__image_size[0] * 0.14

            headers = list(excel_data[0].keys())
            logger.info(("Excel Headers: {0}").format(headers))

            # https://pythoninoffice.com/write-data-to-excel-using-python/
            sheet.cell(self.__first_row, self.__first_col, value = self.__col_json['image'])

            for i in range(0, len(headers), 1):
                sheet.cell( self.__first_row ,i + self.__first_col + 1, value = headers[i])
            
            images=[]
            for i in range(0, len(excel_data), 1):
                d = excel_data[i]
                logger.info(("Excel row: {0}: {1}").format(i, d))
                # Add this because we use "id" display name. It might be the one of the three.
                if d.get('id'):
                    image_path = ('{0}/{1}.jpg').format(img_temp, d['id']) 
                elif d.get('Id'):
                    image_path = ('{0}/{1}.jpg').format(img_temp, d['Id']) 
                elif d.get('ID'):
                    image_path = ('{0}/{1}.jpg').format(img_temp, d['ID']) 

                if os.path.exists(image_path):
                    images.append(image_path)
                    # Load image
                    img = Image(image_path)
                    logger.info(("Load image: {0}").format(image_path))

                    # Add image and align image with cell.
                    # https://www.blog.pythonlibrary.org/2021/08/11/styling-excel-cells-with-openpyxl-and-python/
                    img.width, img.height = self.__image_size
                    cell_tuple = (self.__first_row + i + 1, self.__first_col)
                    cell_str_img = get_column_letter(cell_tuple[1]) + str(cell_tuple[0])
                    sheet.add_image(img, cell_str_img)
                    logger.info(("Add image to excel: {0}").format(cell_str_img))

                    # Set height of the row
                    sheet.row_dimensions[self.__first_row + i + 1].height = self.__image_size[1]  * 0.78

                # Can't delete temp image here. It is required when save excel.
                # os.remove(image_path)
                
                for j in range(0, len(d), 1):
                    cell_tuple_row = (self.__first_row + i + 1, j + self.__first_col + 1)
                    if d[headers[j]] is None:
                        cell_value = ''
                    else:
                        cell_value = d[headers[j]]
                    sheet.cell(cell_tuple_row[0], cell_tuple_row[1], value = cell_value) 
                    cell_str = get_column_letter(cell_tuple_row[1]) + str(cell_tuple_row[0])
                    sheet[cell_str].alignment = Alignment(horizontal="left", vertical="top")

                    logger.info(("Add data to excel: {0}: {1}").format(cell_str, cell_value))
            
            workbook.save(filename=self.__excel_file)
            logger.info(("Excel created: {0}").format(self.__excel_file))

            # Delete temp images
            for i in images:
                if os.path.exists(i):
                    os.remove(i)
                    logger.info(("Image deleted: {0}").format(i))

            # if os.path.exists(img_temp):
            #     os.remove(img_temp)
            #     logger.info(("Image temp folder deleted: {0}").format(img_temp))

        except Exception as e:
            logger.info(("Error: {0}").format(e))

